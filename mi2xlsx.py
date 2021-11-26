# -*- coding: utf-8 -*-
#
# base idea:	http://markkeller.me/2016-09-19-media_info_extractor/


import platform
import subprocess
import os
import sys

from openpyxl import Workbook
from openpyxl import __version__ as opx_v

# Import necessary style classes
from openpyxl.styles import Font, Alignment, Border, Side

print(" 'Python' - Version : " + platform.python_version() + " on os: " + sys.platform)
print(" 'OpenPyXL' - Version : " + opx_v)

# this is a list with items of the type 'string'
# these are extensions of files (in the execution folder) which are 'no media files':
no_media_files = ['.evs', '.py', '.xml', '.pdf', '.docx', '.db', '.exe', '.hide', '.xls', '.xlsx', '.csv', '.log', '.txt']

# here i read an external text file (again with the 'no media file' extensions)
# i believe this is more flexible:
with open ("extensions.txt") as file:
	no_media_files = file.read()
	file.close()
print(no_media_files)

no_media_files = no_media_files.split("\n")
print(no_media_files)

no_media_files = ["." + item for item in no_media_files]
print(no_media_files)

# test - in the following line i filter also 'avi' 'jpg' and 'mpg' files
# no_media_files = ['.evs', '.py', '.xml', '.pdf', '.docx', '.db', '.exe', '.hide', '.xls', '.xlsx', '.csv', '.log', '.avi', '.jpg', '.mpg']

if __name__ == '__main__':

	dir_work = os.getcwd()  # Return a string representing the current working directory
	print('\n', "current working directory:  ", dir_work)

	# path = directory + '/folder'     # os: linux
	# path = directory + '\\folder'  # os: windows
	dir_media = dir_work + '\\media'
	print('\n', "current media directory:  ", dir_media)
	dir_log = dir_work + '\\log'
	print('\n', "current log directory:  ", dir_log)
	dir_xlsx = dir_work + '\\xlsx'
	print('\n', "current workbook directory:  ", dir_xlsx)

	# return a list containing the names of the entries in the directory given by 'path' ,
	# here our 'current working directory'
	# at this time, this 'list_of_all_files' contains all files, without any filter applied
	list_of_all_files = os.listdir(dir_media)
	print("\n", "list of all files:")
	for file in list_of_all_files:
		print("   ", file)

	print("\n")

	only_media_files = []  # here we create an empty list

	# this for-loop starts with the keyword "for" followed by an arbitrary variable name ( in this source: 'file' ),
	# which will hold the values of the following sequence object, which is stepped through.
	# the items of the sequence object are assigned one after the other to the loop variable;
	# to be precise: the variable points to the items.
	# For each item the loop body is executed, 	(we apply the filter (.evs, .py, .xml  ...)
	# to get a list only with our media-files)

	for file in list_of_all_files:

		# os.path.isfile(path) : return 'true' if 'path' is an existing regular file, and 'false' for our 'sub_directory'
		# in this source, at the time of the first loop-step, 'path' is the first 'file' in our 'list_of_all_files'
		#
		# print(os.path.isfile(os.path.join(dir_media,file)))	# prints 'True' or 'False'

		# 'if' is a 'conditional statement' , and in this for-loop we have another for-loop:
		# it starts with 'for' followed by the variable name 'filter'
		# if the item 'file' of 'list_of_files' has no extension listed in our 'filters'-list,
		# than append this 'file' ( = filename ) to our, until now empty, 'only_media_files'

		# if os.path.isfile(file) and all([filter not in file for filter in filters_no_media]):
		# pycharm says we can this above line split in "two if's"

		if os.path.isfile(os.path.join(dir_media,file)):
			if all([filter not in file for filter in no_media_files]):
				only_media_files.append(file)
				# only_media_files.extend(file)

				# list_of_files = only_media_files  # new reference to this list, and this list is 'filtered'
				# print('\n', "list of files (new referenced) : ", "\n", "     ", list_of_files)

	# at this time we have applied our filter rule
	# print('\n', 'only media files (this list is filtered):  ', only_media_files)
	print("filtered list (only media files):")
	for file in only_media_files:
		print("   ", file)

	# for-loop: for every media-file in our 'only_media_files'
	for file in only_media_files:

		# os: linux
		# media_info[file] = subprocess.check_output(['mediainfo.exe %s'%file], shell=True, executable='/bin/bash').split('\n')

		# os: windows
		# media_info[file] = subprocess.check_output(['mediainfo.exe', '%s', '%file'], shell=True, executable='/bin/bash').split('\n')

		# if args is a list, then the first item in this list is considered as the executable and the rest
		# of the items in the list are passed as command line arguments to the program:
		# mi_cmd = ['mediainfo.exe', '-s', (os.path.join(dir_media, file))]

		# in Python 3.6 + you can use the new f - strings:
		# https://docs.python.org/3/whatsnew/3.6.html#pep-498-formatted-string-literals

		# print('\n', "directory:  ", directory, '\n')

		mi_cmd = ['mediainfo ', '--Language=raw ', '--Full ', (os.path.join(dir_media, file)) + ' ',
			f"--Logfile={dir_log}\{file.replace('.', '_')}_raw.log"]

		print('\n', "'mi_cmd' is a list: ", mi_cmd)

		# we get a 'byte string' , therefor we can't 'split'
		# media_info[(os.path.join(dir_media, file)] = subprocess.check_output(mi_cmd, shell=None)

		# we get a 'string', without 'b' ; now we can 'split'
		# print ("This is the 'output' without '.split('\\n')'   :")    # escape slash before the \newline
		# media_info[file] = subprocess.check_output([mi_cmd], universal_newlines=True, encoding='utf-8')
		# print ("-" * 100)
		# print ("media_info '1' is :  ", '\n\n\n', media_info, '\n\n\n')

		# print ("This is the 'output' with '.split('\\n')'   :")

		# media_info is a dictionary:
		# here we generate an empty dictionary
		# the difference between lists and dictionaries:
		# a list is an ordered sequence of objects, whereas dictionaries are unordered sets.
		# but the main difference is, that items in dictionaries are accessed via keys and not via their position.
		media_info = {}

		media_info[file] = subprocess.check_output(mi_cmd, universal_newlines=True, encoding='utf-8').split('\n')

		print ("-" * 100)
		# the answer of 'check_output' is a 'dictionary':
		# the 'media file name' as the key, and a list as the value to this key ,
		# this 'value'-list contains all 'categories' and 'elements' from mediainfo

		# media_info[file] = subprocess.check_output(mi_cmd, shell=None, encoding='utf-8', universal_newlines=True)

		# print("-" * 100)
		# print('media_info is :  ', '\n\n\n', media_info, '\n\n\n')
		# !!!   the output starts with { , followed by the name of our mediafile, test.mp4_hide , in this example ,
		# THAN a lowercase  b  , just before 'General .....
		# this b means in Python3 'byte string': it consists of sequences of 8-bit values, and is for storing to disk,
		# while 'str' consists of sequences of Unicode characters, and is for displaying to humans to read on a computer

		# print ("-" * 100)
		# print('class:  ', type(dict.keys(media_info)),'\n')
		# Get all keys
		# here we see, that 'key' in our 'key-value' pairs in this dictionary is the name of our 'file'
		print("the 'key' :  ", dict.keys(media_info))
		# print("the 'key' :  ", media_info.keys())

		print("-" * 100)
		# Get all values
		# and 'value' is the answer of 'mediainfo.exe'
		print("the 'value' :  ", '\n', dict.values(media_info), '\n')
		# print("the 'value' :  ", '\n', media_info.values(), '\n')

		# print("-" * 100)

		# at this time, we have build up our 'media_info' dictionary for one media file

		# 'dict.items' iterates over the key-value pairs of our dictionary.
		# print("these are the 'items' :   ", '\n', dict.items(media_info), '\n')

		# for-loop: for every 'mediafile' in our 'media_info' -dictionary,
		# we create 'category' - 'append_' and 'category_dictionary' :
		# depart the 'for loop' and 'if - else' with the debugger of your ide, and watch the content of all variables

		for file in media_info.keys():

			# create Workbook object
			wb = Workbook()
			ws = wb.active

			# Create a few styles	examples from realpython
			bold_font = Font(bold=True)
			# big_red_text = Font(color=colors.RED, size=20)			# error
			center_aligned_text = Alignment(horizontal="center")
			double_border_side = Side(border_style="double")
			square_border = Border(top=double_border_side,
                       right=double_border_side,
                       bottom=double_border_side,
                       left=double_border_side)

			# Style some cells!
			# sheet["A2"].font = bold_font
			# sheet["A3"].font = big_red_text
			# sheet["A4"].alignment = center_aligned_text
			# sheet["A5"].border = square_border

			# sheet_row = 2
			sheet_col = 1		# counter for worksheet column
			ws_cat = 0			# counter for worksheet position

			# here starts a loop :
			for line in media_info[file]:

				if line == '':  # there are some empty lines, ignore , don't do anything, go on
					continue
				# print("Line is:  ", line)

				# find the 'category' names (General, Audio, ...)
				# .strip() removes all whitespace at the start and end, including spaces, tabs, newlines and carriage returns
				if ':' not in line.strip():  # there is no colon ':' in the category-lines 'General' - 'Video' - 'Audio' - ....
					print('\n', "No colon ':' in 'line':  ", line)

					category = line.strip()  # update the variable 'category' with the value of stripped 'line' , eg 'General'

					# !=  if values of two operands are not equal, then condition becomes 'true' ,
					# if category != '':  # in the 1st round of this loop 'category' is empty
					# category_list.append(category)
					print('\n', "category is :   ", category, '\n')

					# category_dict = {}  # create a dictionary, or delete content of existing dictionary

					ws_category = wb.create_sheet(category, ws_cat) # insert sheet with name 'category' at 'ws_cat' position

					# set Worksheet tab color.
					ws_category.sheet_properties.tabColor = "FF0000"

					ws_category['A1'] = 'MI Element'	# write text to fixed cell
					ws_category['B1'] = 'MI Value'

					ws_category["A1"].font = bold_font
					ws_category["B1"].font = bold_font

					ws_category["A1"].alignment = center_aligned_text
					ws_category["B1"].alignment = center_aligned_text

					# sheet["A3"].font = big_red_text
					# sheet["A4"].alignment = center_aligned_text

					ws_cat = ws_cat + 1					# insert next worksheet to the next right position

					sheet_row = 2						# on every worksheet the first row is filled with fixed text

				else:
					print("Line is:  ", line)

					# example for 'line' from the audio category ( two ' : ')
					#			Channel positions : Front: L R
					# another example
					#			Tagged date : UTC 2013-12-13 15:39:16
					#
					# second parameter for 'split' is called 'maxsplits', meaning the number of times 'line.split' should do
					mi_element, mi_value = line.split(':', 1)	# first part of splited line is assigned to variable
																# 'mi_element', second part is variable 'mi_value'
					mi_element = mi_element.rstrip()			# delete
					mi_value = mi_value.lstrip()				# delete

					# wb = load_workbook(os.path.join(dir_xlsx, (file.replace('.', '_') + '.xlsx')))
					print(wb.sheetnames)

					ws_category.cell(row = sheet_row, column = sheet_col).value = mi_element
					sheet_col = sheet_col + 1
					ws_category.cell(row = sheet_row, column = sheet_col).value = mi_value
					sheet_row = sheet_row + 1
					sheet_col = sheet_col - 1

					# wb.save(os.path.join(dir_xlsx, (file.replace('.', '_') + '.xlsx')))

				wb.save(os.path.join(dir_xlsx, (file.replace('.', '_') + '.xlsx')))
